# Lee el archvivo historico de excel para generar el json desde ese archivo
# solo se ejecuta para el json original -- NO SE DEBE EJECUTAR MAS

import pandas as pd
from openpyxl import load_workbook

def read_excel_tables(file_path):
    # Load the workbook structure without reading raw data yet
    wb = load_workbook(filename=file_path, data_only=True)
    extracted_tables = {}

    print(f"Workbook '{file_path}' loaded. Sheets found: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
   
        for table in sheet.tables.values():
            # Get the exact cell coordinate string (e.g., 'A1:C10')
            table_name = table.name
            cell_range = table.ref

            start_cell, end_cell = cell_range.split(':')

            # Isolate rows and columns
            start_col = ''.join(filter(str.isalpha, start_cell))
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            end_row = int(''.join(filter(str.isdigit, end_cell)))
            
            # Read data using pandas
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                usecols=f"{start_col}:{ ''.join(filter(str.isalpha, end_cell)) }",
                skiprows=start_row - 1
            )
            
            # Trim the dataframe to match exact table height
            total_rows = end_row - start_row
            df = df.iloc[:total_rows]
            
            extracted_tables[table_name] = df
            print(f"Successfully loaded Table: '{table_name}'")

    return extracted_tables
# 1. Leer el archivo Excel
# Reemplaza 'archivo.xlsx' por el nombre de tu documento
#df = pd.read_excel("Tasas.xlsx", sheet_name="Tasas")

# 2. Guardar el contenido como archivo JSON
# orient='records' crea una lista de objetos JSON por cada fila
#df.to_json('datos.json', orient='records', force_ascii=False, indent=4)

#print('Archivo JSON creado con éxito.')
all_tables = read_excel_tables("Tasas.xlsx")
print(all_tables["Tasas"])
json_data = all_tables["Tasas"].to_json(orient='records', force_ascii=False, indent=4, date_format='iso')
with open('Tasas.json', 'w', encoding='utf-8') as f:
    f.write(json_data)
